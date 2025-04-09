import sys
import os
import pandas as pd
import numpy as np
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout,
                             QPushButton, QWidget, QFileDialog, QLabel,
                             QMessageBox)
from PyQt5.QtCore import Qt
import openpyxl
import xlrd  # 添加对.xls文件的支持
import math
from openpyxl.utils import column_index_from_string


def column_to_index(col_str):
    """Excel列字母转索引（A=0, AA=26等）"""
    return column_index_from_string(col_str) - 1


class ExcelProcessor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.file_path = None
        self.sheet1_data = None
        self.merged_cells_info = None
        self.total_rows = 9999
        self.data_start_row = 14

    def init_ui(self):
        self.setWindowTitle('Excel数据转换工具')
        self.setGeometry(100, 100, 600, 300)

        layout = QVBoxLayout()
        self.label = QLabel('请选择包含Sheet1的Excel文件 (.xls 或 .xlsx)')
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)
        # 添加签名标签到右下角
        self.signature_label = QLabel('by Chase Wang', self)
        self.signature_label.setAlignment(Qt.AlignRight | Qt.AlignBottom)
        self.signature_label.setStyleSheet("color: gray; font-style: italic;")

        self.load_btn = QPushButton('1. 加载Excel文件', self)
        self.load_btn.clicked.connect(self.load_file)
        layout.addWidget(self.load_btn)

        self.process_btn = QPushButton('2. 生成并保存Sheet3', self)  # 修改按钮文本
        self.process_btn.clicked.connect(self.process_and_save_data)
        layout.addWidget(self.process_btn)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def load_file(self):
        """加载Excel文件，支持.xls和.xlsx格式"""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "",
            "Excel Files (*.xlsx *.xls)", options=options)

        if file_path:
            try:
                if not os.access(file_path, os.R_OK):
                    QMessageBox.critical(self, "错误", "无法读取文件，请检查文件权限")
                    return

                self.file_path = file_path

                # 根据文件扩展名选择不同的读取方式
                if file_path.lower().endswith('.xlsx'):
                    self._load_xlsx_file(file_path)
                elif file_path.lower().endswith('.xls'):
                    self._load_xls_file(file_path)
                else:
                    QMessageBox.critical(self, "错误", "不支持的文件格式，请使用.xls或.xlsx文件")
                    return

                self.label.setText(f"已加载文件: {os.path.basename(file_path)}\n准备生成数据表...")

            except Exception as e:
                QMessageBox.critical(self, "错误", f"文件加载失败: {str(e)}")
                self.file_path = None
                self.sheet1_data = None
                self.merged_cells_info = None

    def _load_xlsx_file(self, file_path):
        """加载.xlsx格式文件"""
        # 使用openpyxl加载工作簿以获取合并单元格信息
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb['Sheet1']

        # 存储合并单元格信息
        self.merged_cells_info = {}
        for merge in sheet.merged_cells.ranges:
            top_left_value = sheet.cell(merge.min_row, merge.min_col).value
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    self.merged_cells_info[(row, col)] = top_left_value

        # 使用pandas读取数据
        self.sheet1_data = pd.read_excel(
            file_path,
            sheet_name='Sheet1',
            header=None,
            skiprows=self.data_start_row - 1
        )

    def _load_xls_file(self, file_path):
        """加载.xls格式文件"""
        # 使用xlrd加载工作簿以获取合并单元格信息
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        sheet = wb.sheet_by_name('Sheet1')

        # 存储合并单元格信息
        self.merged_cells_info = {}
        # xlrd的merged_cells返回的是(行索引, 行索引+行数, 列索引, 列索引+列数)
        for rlo, rhi, clo, chi in sheet.merged_cells:
            top_left_value = sheet.cell_value(rlo, clo)
            for row in range(rlo, rhi):
                for col in range(clo, chi):
                    # 转换为1-based索引，与openpyxl一致
                    self.merged_cells_info[(row + 1, col + 1)] = top_left_value

        # 使用pandas读取数据
        self.sheet1_data = pd.read_excel(
            file_path,
            sheet_name='Sheet1',
            header=None,
            skiprows=self.data_start_row - 1,
            engine='xlrd'
        )

    def process_and_save_data(self):
        """处理数据并保存到新文件"""
        if not self.file_path or self.sheet1_data is None:
            QMessageBox.warning(self, "警告", "请先选择有效的Excel文件")
            return

        try:
            # ================== 生成Sheet2 (仅用于计算，不保存) ==================
            sheet2 = pd.DataFrame(index=range(self.total_rows))

            # 添加No序号列
            sheet2['No'] = range(1, self.total_rows + 1)

            # ============== 新的B列和C列计算逻辑 ==============
            b_rows = []
            for row in range(1, self.total_rows + 1):
                if row == 1:
                    b_rows.append(1 + self.data_start_row - 1)
                elif row == 2:
                    b_rows.append(88 + 2 + self.data_start_row - 1)
                else:
                    b_rows.append(85 * (row - 2) + 88 + 2 + self.data_start_row - 1)

            c_rows = []
            for row in range(1, self.total_rows + 1):
                if row == 1:
                    c_rows.append(3 + self.data_start_row - 1)
                elif row == 2:
                    c_rows.append(88 + 4 + self.data_start_row - 1)
                else:
                    c_rows.append(85 * (row - 2) + 88 + 4 + self.data_start_row - 1)

            sheet2['B'] = [self._get_value('H', row) for row in b_rows]
            sheet2['C'] = [self._get_value('H', row) for row in c_rows]

            # ============== 其他列保持原逻辑 ==============
            row_numbers = {
                'D': [self._get_d_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)],
                'E': [self._get_e_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)],
                'G': [self._get_g_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)],
                'H': [self._get_g_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)],
                'K': [self._get_k_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)],
                'L': [self._get_l_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)],
                'U': [self._get_k_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)],
                'V': [self._get_g_row(n) + self.data_start_row - 1 for n in range(1, self.total_rows + 1)]
            }

            sheet2['D'] = [self._get_value('F', row) for row in row_numbers['D']]
            sheet2['E'] = [self._get_value('R', row) for row in row_numbers['E']]
            sheet2['F'] = [f"S{i + 1}" for i in range(self.total_rows)]
            sheet2['G'] = [self._get_value('AB', row) for row in row_numbers['G']]
            sheet2['H'] = [self._get_value('AH', row) for row in row_numbers['H']]
            sheet2['I'] = [f"AC{i + 1}" for i in range(self.total_rows)]
            sheet2['K'] = [self._get_value('K', row, 1000) for row in row_numbers['K']]
            sheet2['L'] = [self._get_value('K', row, 1000) for row in row_numbers['L']]
            sheet2['U'] = [self._get_value('K', row, 1000) for row in row_numbers['U']]
            sheet2['V'] = [
                self._get_value('AD', row, 1000) if self._get_value('AD', row, 1000) > 0 else 1e-30
                for row in row_numbers['V']
            ]

            # 计算列
            sheet2['M'] = np.abs(sheet2['L'] / sheet2['K'].replace(0, np.nan))
            sheet2['N'] = 1 / (60 * 2)
            sheet2['O'] = np.log(sheet2['M'].replace(0, np.nan))
            sheet2['P'] = sheet2['O'] / -sheet2['N']
            sheet2['Q'] = 1 / sheet2['P'].replace(0, np.nan)
            sheet2['R'] = sheet2['Q'] * 2 * math.pi * 60
            sheet2['S'] = np.cos(np.arctan(sheet2['R']))

            sheet2['W'] = np.abs(sheet2['V'] / sheet2['U'].replace(0, np.nan))
            sheet2['X'] = 1 / (60 * 2)
            sheet2['Y'] = np.log(sheet2['W'].replace(0, np.nan))
            sheet2['Z'] = sheet2['Y'] / -sheet2['X']
            sheet2['AA'] = 1 / sheet2['Z'].replace(0, np.nan)
            sheet2['AB'] = sheet2['AA'] * 2 * math.pi * 60
            sheet2['AC'] = np.cos(np.arctan(sheet2['AB']))

            sheet2['F'] = sheet2['S']
            sheet2['I'] = sheet2['AC']
            sheet2 = sheet2[['No'] + [col for col in sheet2.columns if col != 'No']]

            # ================== 生成Sheet3 ==================
            sheet3 = pd.DataFrame()
            sheet3['No'] = range(1, self.total_rows + 1)
            sheet3['C'] = sheet2['B'].copy()

            columns_to_map = {
                'D': 'C',
                'E': 'D',
                'F': 'E',
                'G': 'F',
                'H': 'G',
                'I': 'H',
                'J': 'I'
            }

            for sheet3_col, sheet2_col in columns_to_map.items():
                unique_pairs = sheet2[['B', sheet2_col]].drop_duplicates(subset=['B'])
                mapping = dict(zip(unique_pairs['B'], unique_pairs[sheet2_col]))
                sheet3[sheet3_col] = sheet3['C'].map(mapping)

            sheet3 = sheet3[['No'] + [col for col in sheet3.columns if col != 'No']]

            # ================== 保存结果 ==================
            save_path, _ = QFileDialog.getSaveFileName(
                self, "保存结果文件",
                os.path.join(os.path.dirname(self.file_path), "result.xlsx"),
                "Excel Files (*.xlsx)")

            if not save_path:
                return

            if not save_path.lower().endswith('.xlsx'):
                save_path += '.xlsx'

            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # 只保存Sheet3，不保存Sheet2
                sheet3.to_excel(writer, sheet_name='Sheet3', index=False)

            QMessageBox.information(
                self, "完成",
                f"文件已成功保存到:\n{save_path}\n\n"
                f"原始文件未被修改: {os.path.basename(self.file_path)}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"数据处理失败: {str(e)}")

    # ============== 其他列的行号计算规则 ==============
    def _get_d_row(self, n):
        if n == 1: return 18
        if n == 2: return 88 + 19
        return 85 * (n - 2) + 88 + 19

    def _get_e_row(self, n):
        return self._get_d_row(n)

    def _get_g_row(self, n):
        if n == 1: return 15
        if n == 2: return 88 + 13
        return 85 * (n - 2) + 88 + 13

    def _get_k_row(self, n):
        if n == 1: return 10
        if n == 2: return 88 + 11
        return 85 * (n - 2) + 88 + 11

    def _get_l_row(self, n):
        return self._get_d_row(n)

    def _get_value(self, column, row, multiplier=1):
        try:
            col_idx = column_to_index(column)

            if self.merged_cells_info and (row, col_idx + 1) in self.merged_cells_info:
                value = self.merged_cells_info[(row, col_idx + 1)]
            else:
                if row - self.data_start_row >= len(self.sheet1_data) or row - self.data_start_row < 0:
                    return np.nan
                value = self.sheet1_data.iloc[row - self.data_start_row, col_idx]

            if pd.isna(value):
                return np.nan

            try:
                return float(value) * multiplier
            except (ValueError, TypeError):
                return value

        except Exception as e:
            print(f"获取数据错误 列{column} 行{row}: {str(e)}")
            return np.nan


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    ex = ExcelProcessor()
    ex.show()
    sys.exit(app.exec_())
